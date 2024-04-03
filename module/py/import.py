import sys
import openpyxl
import sqlite3
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import io
import pandas as pd
import numpy as np

filepath = "sample.xlsx"
dbpath = "database.db"

con = sqlite3.connect(dbpath)
# enable the foreign key constrain
con.execute("PRAGMA foreign_keys = ON;")
cur = con.cursor()

cur.execute("DROP TABLE IF EXISTS Quotes")
cur.execute("DROP TABLE IF EXISTS Products;")
cur.execute("DROP TABLE IF EXISTS Users;")

# Users table
cur.execute("CREATE TABLE Users (username TEXT PRIMARY KEY, password TEXT, user_group TEXT);")

# Products table
cur.execute("""CREATE TABLE "Products" (
                        	"Code"	TEXT PRIMARY KEY,
                        	"Barcode"	TEXT,
                        	"Category"	TEXT,
                        	"Description"	TEXT,
                        	"Color"	TEXT,
                        	"Range"	TEXT,
                        	"Price"	REAL,
                        	"Qty"	REAL,
                        	"WELS"	TEXT,
                        	"REG"	TEXT,
                        	"Rating"	REAL,
                        	"WaterCon"	REAL,
                        	"Warranty"	TEXT,
                        	"img"	REAL
                            );"""
            )
con.commit()

# import product data
df = pd.read_excel(filepath, engine="openpyxl")
df_clean = df.dropna(axis=1, how='all').dropna(how='all')
df_clean = df_clean.rename(columns={"Book Price (Ex. GST)": "Price", "Qty/Box": "Qty", "Color Finish": "Color", "WELS LIC.": "WELS", "REG. Number": "REG", "STAR Rating": "Rating", "Water Cconsump.": "WaterCon"})
df_clean = df_clean.round({"Price": 2})

# Data wraggling and Transformation
df_clean["WELS"].replace('NOT REQUIRE', np.nan, inplace=True)
df_clean["Category"].replace(np.nan, "OTHER", inplace=True)
df_clean["Color"].replace("BRUSHED GOLD", "BRUSH GOLD", inplace=True)
df_clean["Color"].replace("BRUSHED GUN METAL", "BRUSH GUN METAL", inplace=True)
df_clean["Color"] = df_clean["Color"].str.upper()

# add column called img
df_clean["img"] = np.nan
# import data to sqlite
df_clean.to_sql("Products", con, if_exists='append', index=False)
con.commit()

# import product image
wb = openpyxl.load_workbook(filepath)
ws = wb[wb.sheetnames[0]]
image_loader = SheetImageLoader(ws)
img_locs = []
for image_loc in image_loader._images:
    img_locs.append(image_loc)

# clean the image_locs and sort
img_locs_filtered = [item for item in img_locs if item.startswith('A')]
img_locs_ft = sorted(img_locs_filtered, key=lambda x: int(x[1:]))

# Extract images to BLOB
imgs_blob = []
for cell in img_locs_ft:
    img = image_loader.get(cell)
    stream = io.BytesIO()
    img.save(stream, format="PNG")
    imgs_blob.append(stream.getvalue())

# Extract Product Labels
img_labels = []
for label, img_blob in zip(img_locs_ft, imgs_blob):
    img_labels.append((img_blob, ws[label.replace("A", "B")].value))

# import product image to sqlite
cur.executemany("UPDATE Products SET img = ? WHERE Code = ?", img_labels)

# User Quote Recording
cur.execute("""CREATE TABLE Quotes (username TEXT,
                                    product TEXT,
                                    price REAL,
                                    quantity INTEGER,
                                    quote_group INTEGER CHECK (typeof(quote_group)='integer' AND quote_group > 0 AND quote_group <= 5),
                                    FOREIGN KEY (username) REFERENCES Users (username),
                                    FOREIGN KEY (product) REFERENCES Products (Code)
                                    );""")

con.commit()
con.close()

print("finish")
# sys.stdout.flush()
